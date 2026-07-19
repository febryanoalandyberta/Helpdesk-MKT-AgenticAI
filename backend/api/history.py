import logging
import io
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy import desc, text
from database import get_db
from models.device import Device
from models.telemetry_history import TelemetryLog
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

router = APIRouter()

@router.get("/{device_id}/history")
async def get_device_history(
    device_id: str,
    page: int = 1,
    limit: int = 100,
    start_date: str = None,
    end_date: str = None,
    db: AsyncSession = Depends(get_db)
):
    """
    Get paginated telemetry history for a specific device.
    Supports optional date filtering (YYYY-MM-DD).
    """
    q = await db.execute(select(Device).where(Device.device_id == device_id))
    device = q.scalar_one_or_none()
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")

    query = select(TelemetryLog).where(TelemetryLog.device_id == device_id)

    if start_date:
        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            query = query.where(TelemetryLog.time >= start_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid start_date format. Use YYYY-MM-DD")
            
    if end_date:
        try:
            end_dt = datetime.strptime(end_date, "%Y-%m-%d")
            # Set to end of day
            end_dt = end_dt.replace(hour=23, minute=59, second=59)
            query = query.where(TelemetryLog.time <= end_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid end_date format. Use YYYY-MM-DD")

    # Pagination
    offset = (page - 1) * limit
    query = query.order_by(desc(TelemetryLog.time)).offset(offset).limit(limit)

    result = await db.execute(query)
    logs = result.scalars().all()

    return {
        "device_id": device_id,
        "device_name": device.device_name,
        "page": page,
        "limit": limit,
        "data": [log.to_dict() for log in logs]
    }


@router.get("/{device_id}/history/export")
async def export_device_history_excel(
    device_id: str,
    start_date: str = None,
    end_date: str = None,
    db: AsyncSession = Depends(get_db)
):
    """
    Export telemetry history to a styled Excel (.xlsx) file.
    """
    q = await db.execute(select(Device).where(Device.device_id == device_id))
    device = q.scalar_one_or_none()
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")

    query = select(TelemetryLog).where(TelemetryLog.device_id == device_id).order_by(desc(TelemetryLog.time))
    
    if start_date:
        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            query = query.where(TelemetryLog.time >= start_dt)
        except ValueError:
            pass
            
    if end_date:
        try:
            end_dt = datetime.strptime(end_date, "%Y-%m-%d")
            end_dt = end_dt.replace(hour=23, minute=59, second=59)
            query = query.where(TelemetryLog.time <= end_dt)
        except ValueError:
            pass

    # Limit to 50,000 rows to prevent massive memory usage on export
    query = query.limit(50000)
    result = await db.execute(query)
    logs = result.scalars().all()

    # Create Excel Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Telemetry Log"

    # Define Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    align_center = Alignment(horizontal="center", vertical="center")

    headers = ["Timestamp", "Device Name", "CPU Usage (%)", "RAM Usage (%)", "Temperature (°C)", "Active App", "Active URL"]
    
    # Write Headers
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = align_center

    # Write Data
    from zoneinfo import ZoneInfo
    jkt_tz = ZoneInfo("Asia/Jakarta")

    for row_num, log in enumerate(logs, 2):
        local_time_str = ""
        if log.time:
            # log.time is naive UTC from DB, make it timezone-aware UTC, then convert to Asia/Jakarta
            utc_time = log.time.replace(tzinfo=ZoneInfo("UTC"))
            local_time = utc_time.astimezone(jkt_tz)
            local_time_str = local_time.strftime("%Y-%m-%d %H:%M:%S")

        row_data = [
            local_time_str,
            device.device_name,
            f"{log.cpu_usage:.1f}%" if log.cpu_usage is not None else "",
            f"{log.ram_usage:.1f}%" if log.ram_usage is not None else "",
            f"{log.temperature:.1f}°C" if log.temperature is not None else "",
            log.active_app or "",
            log.active_url or ""
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            if col_num in [1, 3, 4, 5]:  # Center align timestamp and metrics
                cell.alignment = align_center

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        # Cap max width to 50 for URLs so it doesn't get ridiculously wide
        if adjusted_width > 50:
            adjusted_width = 50
        ws.column_dimensions[column].width = adjusted_width

    # Save to memory stream
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename_date = start_date if start_date else datetime.now().strftime("%Y-%m-%d")
    filename = f"TelemetryLog_{device.device_name}_{filename_date}.xlsx"
    
    headers = {
        'Content-Disposition': f'attachment; filename="{filename}"'
    }
    
    return StreamingResponse(
        output, 
        headers=headers, 
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
