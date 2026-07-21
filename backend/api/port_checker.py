"""
Port Checker API — Manages hardware physical connection incidents
"""
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from fastapi import APIRouter, Depends, HTTPException, Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, desc, update
from pydantic import BaseModel
from typing import Optional, List
from database import get_db
from models.port_checker import PortCheckerLog

router = APIRouter(prefix="/api/port-checker", tags=["Port Checker"])

class CreatePortLogRequest(BaseModel):
    summary: str
    category: Optional[str] = None
    hardware_type: Optional[str] = None
    hardware_name: Optional[str] = None
    device_name: Optional[str] = None
    device_id: Optional[str] = None
    site_name: Optional[str] = None # Not in PortCheckerLog model currently, but kept for compatibility with incoming payloads

@router.post("/")
async def create_port_log(data: CreatePortLogRequest, db: AsyncSession = Depends(get_db)):
    # Fallback to map the device_id if not present in request.
    mapped_device_id = data.device_id

    if not mapped_device_id and data.device_name:
        # Search for device by device_name (hostname)
        from models.device import Device
        q = await db.execute(select(Device.device_id).where(Device.device_name == data.device_name))
        found_id = q.scalar_one_or_none()
        if found_id:
            mapped_device_id = str(found_id)

    log_data = {
        "summary": data.summary,
        "category": data.category,
        "device_name": data.device_name,
        "hardware_type": data.hardware_type,
        "hardware_name": data.hardware_name,
        "device_id": mapped_device_id,
        "is_read": False
    }
    
    port_log = PortCheckerLog(**log_data)
    db.add(port_log)
    await db.commit()
    await db.refresh(port_log)
    return port_log.to_dict()

@router.get("/{device_id}")
async def get_device_port_logs(
    device_id: str, 
    limit: int = 50, 
    db: AsyncSession = Depends(get_db)
):
    # For a specific device, fetch logs. If device_id is unknown, we can also query by device_name
    query = select(PortCheckerLog).where(PortCheckerLog.device_id == device_id).order_by(desc(PortCheckerLog.created_at)).limit(limit)
    q = await db.execute(query)
    logs = q.scalars().all()
    
    # If no logs by device_id, try by device_name (for backward compatibility if the agent only sent device_name)
    if not logs:
        # We need the device_name from devices table if we had it, but since we just have device_id here, 
        # let's assume the frontend will only query by device_id.
        pass

    return {"logs": [l.to_dict() for l in logs]}

@router.post("/mark-read/{device_id}")
async def mark_device_logs_read(device_id: str, db: AsyncSession = Depends(get_db)):
    # Update all unread logs for this device to read
    stmt = update(PortCheckerLog).where(
        (PortCheckerLog.device_id == device_id) & (PortCheckerLog.is_read == False)
    ).values(is_read=True)
    await db.execute(stmt)
    await db.commit()
    return {"success": True}

@router.get("/export/{device_id}")
async def export_port_logs(device_id: str, db: AsyncSession = Depends(get_db)):
    query = select(PortCheckerLog).where(PortCheckerLog.device_id == device_id).order_by(desc(PortCheckerLog.created_at))
    q = await db.execute(query)
    logs = q.scalars().all()

    if not logs:
        raise HTTPException(status_code=404, detail="No port checker history found for this device")

    # Create Excel Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Port Checker History"

    # Define Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    align_center = Alignment(horizontal="center", vertical="center")

    headers = ["Waktu (WIB)", "Kategori", "Perangkat Terkait", "Ringkasan", "Status"]
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = align_center

    from zoneinfo import ZoneInfo
    jkt_tz = ZoneInfo("Asia/Jakarta")

    for row_num, l in enumerate(logs, 2):
        local_time_str = ""
        if l.created_at:
            utc_time = l.created_at.replace(tzinfo=ZoneInfo("UTC"))
            local_time = utc_time.astimezone(jkt_tz)
            local_time_str = local_time.strftime("%Y-%m-%d %H:%M:%S")

        hardware_desc = f"{l.hardware_type or ''} - {l.hardware_name or ''}".strip(' -')
        
        row_data = [
            local_time_str,
            l.category or "-",
            hardware_desc if hardware_desc else "-",
            l.summary or "-",
            "Sudah Dibaca" if l.is_read else "Belum Dibaca"
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            if col_num in [1, 2, 5]:  # Center align timestamp, category, status
                cell.alignment = align_center

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        if adjusted_width > 50:
            adjusted_width = 50
        ws.column_dimensions[column].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    headers_dict = {
        'Content-Disposition': f'attachment; filename="port_checker_{device_id}.xlsx"'
    }
    return Response(content=output.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers_dict)
